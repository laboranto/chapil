from fastapi import FastAPI, HTTPException, UploadFile, File
from PIL import Image, ImageOps
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from contextlib import asynccontextmanager
from pydantic import BaseModel, field_validator
from typing import Optional, List
from datetime import date, timedelta, datetime
from pathlib import Path
from enum import Enum
import io
import os
import json, redis
import time
import openpyxl
from database import get_db, init_db

# redis 접속 선언
redis_client = redis.Redis(host='localhost', port=6379, db=0)

BASE_DIR = Path(__file__).parent

# STATIC_DIR: 로컬에서는 ../static, Docker에서는 ENV STATIC_DIR=/app/static 으로 주입된다.
STATIC_DIR = Path(os.environ.get("STATIC_DIR", str(BASE_DIR.parent / "static")))

# BACKUP_DIR: 자동 백업 저장 경로. Docker에서는 ENV BACKUP_DIR=/data/backups 으로 주입한다.
BACKUP_DIR = Path(os.environ.get("BACKUP_DIR", str(BASE_DIR.parent / "backups")))
BACKUP_MAX_COUNT = int(os.environ.get("BACKUP_MAX_COUNT", "30"))

DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR.parent / "data")))
IMAGE_PATH          = DATA_DIR / "car_image.jpg"
IMAGE_ORIGINAL_PATH = DATA_DIR / "car_image_original.jpg"
MAX_IMAGE_BYTES = 5 * 1024 * 1024

# DIST_DIR: React 빌드 결과물 경로. Docker 빌드 시에만 존재한다.
DIST_DIR = BASE_DIR / "frontend" / "dist"


# ── 백업 ──────────────────────────────────────────────────────────────

def build_export_data() -> dict:
    conn = get_db()
    settings_rows = conn.execute("SELECT key, value FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in settings_rows}
    fuel_rows = conn.execute(
        "SELECT date, type, amount, unit_price, liters, odometer, memo FROM fuel ORDER BY date ASC, id ASC"
    ).fetchall()
    maintenance_rows = conn.execute(
        "SELECT date, item, amount, odometer, memo FROM maintenance ORDER BY date ASC, id ASC"
    ).fetchall()
    other_rows = conn.execute(
        "SELECT date, item, amount, odometer, memo FROM other ORDER BY date ASC, id ASC"
    ).fetchall()
    conn.close()
    return {
        "vehicle": {
            "car_birth": settings.get("car_birth", ""),
            "car_type":  settings.get("car_type", ""),
            "car_brand": settings.get("car_brand", ""),
            "car_model": settings.get("car_model", ""),
            "car_plate": settings.get("car_plate", ""),
            "car_fuel":  settings.get("car_fuel", ""),
        },
        "fuel":        [dict(r) for r in fuel_rows],
        "maintenance": [dict(r) for r in maintenance_rows],
        "other":       [dict(r) for r in other_rows],
    }


def backup_if_needed():
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    existing = sorted(BACKUP_DIR.glob("chapil_backup_*.json"), key=lambda p: p.stat().st_mtime)

    # 24시간 이내 백업이 존재하면 건너뜀
    if existing and (time.time() - existing[-1].stat().st_mtime) < 86400:
        return

    filename = f"chapil_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    (BACKUP_DIR / filename).write_text(
        json.dumps(build_export_data(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # 초과분 삭제 (오래된 것부터)
    existing = sorted(BACKUP_DIR.glob("chapil_backup_*.json"), key=lambda p: p.stat().st_mtime)
    for old in existing[:-BACKUP_MAX_COUNT]:
        old.unlink()


# ── lifespan: 앱 시작/종료 시점에 실행될 코드를 정의 ─────────────────
# 기존 @app.on_event("startup") 방식은 deprecated 되었고,
# 현재 FastAPI 권장 방식은 contextlib.asynccontextmanager를 사용한 lifespan이다.
# yield 이전: 앱 시작 시 실행 / yield 이후: 앱 종료 시 실행
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    seed_demo()
    backup_if_needed()
    yield


app = FastAPI(title="차계부", lifespan=lifespan)

# 정적 파일 서빙 — Docker 빌드 시에만 static/ 폴더가 존재한다.
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# CORS 미들웨어: React SPA가 다른 포트(예: localhost:3000)에서 이 API를 호출할 수 있도록 허용.
# 배포 시에는 allow_origins를 실제 도메인으로 제한할 것.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# 정비/기타 항목 선택지 — 프론트엔드 폼에서 드롭다운으로 사용
MAINTENANCE_ITEMS = [
    "엔진오일 및 오일필터",
    "에어클리너 필터",
    "에어컨필터(항균필터)",
    "점화플러그",
    "브레이크 패드 (앞)",
    "브레이크 패드 (뒤)",
    "브레이크 오일",
    "타이어",
    "타이어 위치",
    "와이퍼 블레이드",
    "엔진부동액(냉각수)",
    "에어컨(냉매, 콤프레서, 콘덴서 등)",
    "배터리",
    "점검",
    "일반수리",
    "기타",
]

OTHER_ITEMS = [
    "자동차 보험",
    "자동차검사",
    "세차",
    "주차비",
    "통행료",
    "실외용품",
    "실내용품",
    "교통범칙금",
    "기타",
]


# ── 차량 정보 Enum ────────────────────────────────────────────────────

class CarType(str, Enum):
    MICROCAR = "microcar"    # 경차
    SMALL    = "small"       # 소형차
    COMPACT  = "compact"     # 준중형차
    MIDSIZE  = "midsize"     # 중형차
    LARGE    = "large"       # 준대형차
    FULLSIZE = "fullsize"    # 대형차
    SUV      = "suv"         # SUV
    RV       = "rv"          # RV
    VAN      = "van"         # 밴
    TRUCK    = "truck"       # 트럭
    OTHER    = "other"       # 기타


class CarFuel(str, Enum):
    # 주유 계열 — 백엔드 로직 동일, 프론트 용어 "주유", 단위 L
    GASOLINE = "gasoline"    # 휘발유
    DIESEL   = "diesel"      # 경유
    HEV      = "hev"         # 하이브리드 (자동충전, 주유만)
    # 충전 계열 — 백엔드 로직 동일, 프론트 용어 "충전", 단위 상이
    LPG      = "lpg"         # LPG (단위: L)
    EV       = "ev"          # 전기 (단위: kWh, 효율명: 전비)
    FCEV     = "fcev"        # 수소전지 (단위: kg)
    # TODO: 복합 로직 필요 — 추후 별도 구현
    BIFUEL   = "bifuel"      # 바이퓨얼 (휘발유+LPG) — 기록 분리 필요
    PHEV     = "phev"        # 플러그인 하이브리드 — 주유+충전 병존


# 선택지 목록 — GET /api/settings/options 에서 반환
# ui_term: 프론트에서 '주유 기록' 탭/버튼에 표시할 용어
# qty_unit: 수량 입력 단위
# price_unit: 단가 단위
# economy_unit: 효율 단위
# economy_label: 효율 명칭 (연비 vs 전비)
CAR_TYPE_OPTIONS = [
    {"code": "microcar",  "label": "경차"},
    {"code": "small",     "label": "소형차"},
    {"code": "compact",   "label": "준중형차"},
    {"code": "midsize",   "label": "중형차"},
    {"code": "large",     "label": "준대형차"},
    {"code": "fullsize",  "label": "대형차"},
    {"code": "suv",       "label": "SUV"},
    {"code": "rv",        "label": "RV"},
    {"code": "van",       "label": "밴"},
    {"code": "truck",     "label": "트럭"},
    {"code": "other",     "label": "기타"},
]

CAR_FUEL_OPTIONS = [
    {"code": "gasoline", "label": "휘발유",           "ui_term": "주유", "qty_unit": "L",   "price_unit": "원/L",   "economy_unit": "km/L",   "economy_label": "연비"},
    {"code": "diesel",   "label": "경유",             "ui_term": "주유", "qty_unit": "L",   "price_unit": "원/L",   "economy_unit": "km/L",   "economy_label": "연비"},
    {"code": "hev",      "label": "하이브리드",        "ui_term": "주유", "qty_unit": "L",   "price_unit": "원/L",   "economy_unit": "km/L",   "economy_label": "연비"},
    {"code": "lpg",      "label": "LPG",              "ui_term": "충전", "qty_unit": "L",   "price_unit": "원/L",   "economy_unit": "km/L",   "economy_label": "연비"},
    {"code": "ev",       "label": "전기",             "ui_term": "충전", "qty_unit": "kWh", "price_unit": "원/kWh", "economy_unit": "km/kWh", "economy_label": "전비"},
    {"code": "fcev",     "label": "수소전지",          "ui_term": "충전", "qty_unit": "kg",  "price_unit": "원/kg",  "economy_unit": "km/kg",  "economy_label": "연비"},
    {"code": "bifuel",   "label": "바이퓨얼 (휘발유+LPG)", "ui_term": "주유", "qty_unit": "L", "price_unit": "원/L", "economy_unit": "km/L",   "economy_label": "연비"},  # TODO
    {"code": "phev",     "label": "플러그인 하이브리드", "ui_term": "주유", "qty_unit": "L",  "price_unit": "원/L",   "economy_unit": "km/L",   "economy_label": "연비"},  # TODO
]

_CAR_TYPE_BY_LABEL = {opt["label"]: opt["code"] for opt in CAR_TYPE_OPTIONS}
_CAR_FUEL_BY_LABEL = {opt["label"]: opt["code"] for opt in CAR_FUEL_OPTIONS}


# ── 데이터 가져오기/내보내기 상수 ─────────────────────────────────────────

# LLM에게 전달할 변환 프롬프트. 출력 포맷을 명시적으로 고정한다.
IMPORT_PROMPT = """\
아래는 차계부 앱 "차필"의 데이터 가져오기 형식입니다.
첨부한 파일(기존 차계부 앱에서 내보낸 파일)과 예제 파일(chapil_template.json)을 참고해서,
기존 데이터를 아래 JSON 형식으로 변환해 주세요.
변환 결과는 코드 블록(```json) 안에 출력하거나, JSON 파일로 제공해 주세요.

[출력 형식]
{
  "vehicle": { 차량 기본 정보 },
  "fuel": [ 주유 기록 배열 ],
  "maintenance": [ 정비 기록 배열 ],
  "other": [ 기타 지출 배열 ]
}

[vehicle - 차량 기본 정보 필드]
• car_type: 차종 (경차/소형차/준중형차/중형차/준대형차/대형차/SUV/RV/밴/트럭/기타)
• car_brand: 브랜드, 문자열 (예: 현대, 기아, BMW)
• car_model: 차량 이름, 문자열 (예: 모닝, 코나, 아이오닉5)
• car_plate: 차량등록번호, 문자열 (예: 123가4567)
• car_birth: 출고일자, YYYY-MM-DD 형식
• car_fuel: 연료 종류 (휘발유/경유/LPG/바이퓨얼/전기/플러그인하이브리드/수소전지)

[fuel - 주유 기록 필드]
• date: 날짜, YYYY-MM-DD 형식 (필수)
• type: "가득주유" 또는 "부분주유" (기본값: "가득주유")
• amount: 주유 금액, 원 단위 정수 (필수)
• unit_price: 리터당 단가, 원 단위 정수, 없으면 null
• liters: 주유량, 리터 단위 실수, 없으면 null
• odometer: 누적 주행거리, km 단위 정수 (필수)
• memo: 메모 문자열, 없으면 ""

[maintenance - 정비 기록 필드]
• date: 날짜, YYYY-MM-DD 형식 (필수)
• item: 정비 항목 이름, 문자열 (필수)
• amount: 정비 비용, 원 단위 정수, 없으면 0
• odometer: 주행거리, km 단위 정수 (필수)
• memo: 메모 문자열, 없으면 ""

[other - 기타 지출 필드]
• date: 날짜, YYYY-MM-DD 형식 (필수)
• item: 지출 항목 이름, 문자열 (필수)
• amount: 금액, 원 단위 정수, 없으면 0
• odometer: 주행거리, km 단위 정수, 없으면 null
• memo: 메모 문자열, 없으면 ""

[변환 규칙]
• 날짜는 반드시 YYYY-MM-DD 형식으로 변환할 것
• 금액·거리의 콤마, 단위 문자(원, km, L 등)는 제거하고 숫자만 남길 것
• 해당 카테고리에 데이터가 없으면 빈 배열 []로 출력할 것
• id, interval_km, fuel_economy, created_at 필드는 출력하지 않아도 됨\
"""

# 예제 JSON — 실제 DB 스키마 기반으로 생성
IMPORT_TEMPLATE = {
    "vehicle": {
        "car_type": "중형차",
        "car_brand": "현대",
        "car_model": "아반떼",
        "car_plate": "123가4567",
        "car_birth": "2020-02-22",
        "car_fuel": "휘발유"
    },
    "fuel": [
        {
            "date": "2024-01-15",
            "type": "가득주유",
            "amount": 80000,
            "unit_price": 1750,
            "liters": 45.7,
            "odometer": 12500,
            "memo": ""
        },
        {
            "date": "2024-01-28",
            "type": "부분주유",
            "amount": 30000,
            "unit_price": 1760,
            "liters": None,
            "odometer": 12800,
            "memo": "고속도로 휴게소"
        }
    ],
    "maintenance": [
        {
            "date": "2024-01-10",
            "item": "엔진오일 및 오일필터",
            "amount": 85000,
            "odometer": 12000,
            "memo": ""
        }
    ],
    "other": [
        {
            "date": "2024-01-05",
            "item": "자동차 보험",
            "amount": 250000,
            "odometer": None,
            "memo": "1년 갱신"
        }
    ]
}


# ── Pydantic 요청 바디 모델 ───────────────────────────────────────────
# Pydantic 모델은 POST/PUT 요청의 JSON 바디를 자동으로 파싱하고 타입을 검증한다.
# 기존의 Form(...) 방식을 대체한다.
# Optional[타입] = None : 값이 없어도 되는 필드

class FuelBody(BaseModel):
    date: str
    type: str = "가득주유"
    amount: int
    unit_price: Optional[int] = None
    liters: Optional[float] = None
    odometer: int
    location: Optional[str] = None
    memo: Optional[str] = None


class MaintenanceBody(BaseModel):
    date: str
    item: str
    amount: int = 0
    odometer: int
    location: Optional[str] = None
    memo: Optional[str] = None


class OtherBody(BaseModel):
    date: str
    item: str
    amount: int = 0
    odometer: Optional[int] = None
    location: Optional[str] = None
    memo: Optional[str] = None


class SettingsBody(BaseModel):
    car_birth: str = ""
    car_type: Optional[CarType] = None
    car_brand: str = ""
    car_model: str = ""
    car_plate: str = ""
    car_fuel: Optional[CarFuel] = None

    @field_validator("car_birth")
    @classmethod
    def validate_car_birth(cls, v: str) -> str:
        if v:
            try:
                date.fromisoformat(v)
            except ValueError:
                raise ValueError("날짜 형식은 YYYY-MM-DD여야 합니다.")
        return v


class ImportRecordFuel(BaseModel):
    date: str
    type: str = "가득주유"
    amount: int
    unit_price: Optional[int] = None
    liters: Optional[float] = None
    odometer: int
    location: Optional[str] = None
    memo: str = ""


class ImportRecordMaintenance(BaseModel):
    date: str
    item: str
    amount: int = 0
    odometer: int
    location: Optional[str] = None
    memo: str = ""


class ImportRecordOther(BaseModel):
    date: str
    item: str
    amount: int = 0
    odometer: Optional[int] = None
    location: Optional[str] = None
    memo: str = ""


class ImportVehicle(BaseModel):
    car_type: str = ""
    car_brand: str = ""
    car_model: str = ""
    car_plate: str = ""
    car_birth: str = ""
    car_fuel: str = ""

    @field_validator("car_type")
    @classmethod
    def normalize_car_type(cls, v: str) -> str:
        return _CAR_TYPE_BY_LABEL.get(v, v)

    @field_validator("car_fuel")
    @classmethod
    def normalize_car_fuel(cls, v: str) -> str:
        return _CAR_FUEL_BY_LABEL.get(v, v)


class ImportBody(BaseModel):
    vehicle: Optional[ImportVehicle] = None
    fuel: List[ImportRecordFuel] = []
    maintenance: List[ImportRecordMaintenance] = []
    other: List[ImportRecordOther] = []


# ── 대시보드 ──────────────────────────────────────────────────────────
# 메인 화면에 필요한 요약 데이터를 한 번에 반환한다.
@app.get("/api/dashboard")
def get_dashboard():
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key='car_birth'").fetchone()
    car_birth_val = row["value"] if row else ""
    car_birth = date.fromisoformat(car_birth_val) if car_birth_val else None
    total_days = (date.today() - car_birth).days if car_birth else None

    recent_fuel = conn.execute(
        "SELECT * FROM fuel ORDER BY date DESC, id DESC LIMIT 5"
    ).fetchall()

    last_maintenance = conn.execute(
        "SELECT * FROM maintenance ORDER BY date DESC, id DESC LIMIT 1"
    ).fetchone()

    last_other = conn.execute(
        "SELECT * FROM other ORDER BY date DESC, id DESC LIMIT 1"
    ).fetchone()

    cutoff = (date.today() - timedelta(days=30)).isoformat()
    fuel_30d  = conn.execute("SELECT SUM(amount) as total FROM fuel WHERE date >= ?", (cutoff,)).fetchone()["total"] or 0
    maint_30d = conn.execute("SELECT SUM(amount) as total FROM maintenance WHERE date >= ?", (cutoff,)).fetchone()["total"] or 0
    other_30d = conn.execute("SELECT SUM(amount) as total FROM other WHERE date >= ?", (cutoff,)).fetchone()["total"] or 0

    avg_economy = conn.execute(
        "SELECT AVG(fuel_economy) as avg FROM fuel"
        " WHERE fuel_economy IS NOT NULL AND fuel_economy > 0 AND fuel_economy <= 50"
        " AND odometer > 0 AND (interval_km IS NULL OR interval_km < odometer * 0.95)"
    ).fetchone()["avg"]

    latest_odometer_row = conn.execute("""
        SELECT odometer FROM (
            SELECT date, odometer FROM fuel WHERE odometer IS NOT NULL
            UNION ALL
            SELECT date, odometer FROM maintenance WHERE odometer IS NOT NULL
            UNION ALL
            SELECT date, odometer FROM other WHERE odometer IS NOT NULL
        ) ORDER BY date DESC, odometer DESC LIMIT 1
    """).fetchone()

    conn.close()
    return {
        "car_birth": str(car_birth) if car_birth else "",
        "total_days": total_days,
        # sqlite3.Row는 JSON 직렬화가 안 되므로 dict()로 변환
        "recent_fuel": [dict(r) for r in recent_fuel],
        "last_maintenance": dict(last_maintenance) if last_maintenance else None,
        "last_other": dict(last_other) if last_other else None,
        "cost_last_30d": fuel_30d + maint_30d + other_30d,
        "avg_economy": round(avg_economy, 2) if avg_economy else None,
        "latest_odometer": latest_odometer_row["odometer"] if latest_odometer_row else None,
    }


# ── 설정 ──────────────────────────────────────────────────────────────

# 차량 정보 선택지 목록 반환 — 프론트엔드 드롭다운에서 사용
@app.get("/api/settings/options")
def settings_options():
    return {
        "car_type": CAR_TYPE_OPTIONS,
        "car_fuel": CAR_FUEL_OPTIONS,
    }


@app.get("/api/settings")
def get_settings():
    conn = get_db()
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    conn.close()
    data = {r["key"]: r["value"] for r in rows}
    car_type_raw = data.get("car_type", "")
    car_fuel_raw = data.get("car_fuel", "")
    return {
        "car_birth":  data.get("car_birth", ""),
        "car_type":   _CAR_TYPE_BY_LABEL.get(car_type_raw, car_type_raw),
        "car_brand":  data.get("car_brand", ""),
        "car_model":  data.get("car_model", ""),
        "car_plate":  data.get("car_plate", ""),
        "car_fuel":   _CAR_FUEL_BY_LABEL.get(car_fuel_raw, car_fuel_raw),
    }


@app.put("/api/settings")
def update_settings(body: SettingsBody):
    conn = get_db()
    fields = {
        "car_birth":  body.car_birth,
        "car_type":   body.car_type.value if body.car_type else "",
        "car_brand":  body.car_brand,
        "car_model":  body.car_model,
        "car_plate":  body.car_plate,
        "car_fuel":   body.car_fuel.value if body.car_fuel else "",
    }
    for key, value in fields.items():
        conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            (key, value),
        )
    conn.commit()
    conn.close()
    return fields


# ── 차량 이미지 ───────────────────────────────────────────────────────

@app.post("/api/settings/image/original", status_code=201)
async def upload_car_image_original(file: UploadFile = File(...)):
    content = await file.read()
    if len(content) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=413, detail="이미지 파일은 5MB 이하여야 합니다.")
    try:
        img = Image.open(io.BytesIO(content))
        img = ImageOps.exif_transpose(img)
        img = img.convert("RGB")
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        img.save(IMAGE_ORIGINAL_PATH, "JPEG", quality=95, optimize=True)
    except Exception:
        raise HTTPException(status_code=400, detail="이미지 파일을 처리할 수 없습니다.")
    return {"ok": True}


@app.get("/api/settings/image/original")
def get_car_image_original():
    if not IMAGE_ORIGINAL_PATH.exists():
        raise HTTPException(status_code=404, detail="원본 이미지가 없습니다.")
    return FileResponse(IMAGE_ORIGINAL_PATH, media_type="image/jpeg")


@app.post("/api/settings/image", status_code=201)
async def upload_car_image(file: UploadFile = File(...)):
    content = await file.read()
    if len(content) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=413, detail="이미지 파일은 5MB 이하여야 합니다.")
    try:
        img = Image.open(io.BytesIO(content))
        img = ImageOps.exif_transpose(img)
        img = img.convert("RGB")
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        img.save(IMAGE_PATH, "JPEG", quality=85, optimize=True)
    except Exception:
        raise HTTPException(status_code=400, detail="이미지 파일을 처리할 수 없습니다.")
    return {"ok": True}


@app.get("/api/settings/image")
def get_car_image():
    if not IMAGE_PATH.exists():
        raise HTTPException(status_code=404, detail="이미지가 없습니다.")
    return FileResponse(IMAGE_PATH, media_type="image/jpeg")


@app.delete("/api/settings/image", status_code=204)
def delete_car_image():
    if IMAGE_PATH.exists():
        IMAGE_PATH.unlink()
    if IMAGE_ORIGINAL_PATH.exists():
        IMAGE_ORIGINAL_PATH.unlink()


# ── 데이터 가져오기 / 내보내기 ─────────────────────────────────────────

# @app.get("/api/import/template")
# def import_template():
#     """예제 JSON 파일 다운로드 — LLM에게 차필 형식을 알려주기 위한 샘플"""
#     content = json.dumps(IMPORT_TEMPLATE, ensure_ascii=False, indent=2)
#     return Response(
#         content=content,
#         media_type="application/json",
#         headers={"Content-Disposition": "attachment; filename=chapil_template.json"},
#     )


# @app.get("/api/import/prompt")
# def import_prompt():
#     """LLM에게 전달할 변환 프롬프트 반환"""
#     return {"prompt": IMPORT_PROMPT}


# ── 마이클 xlsx 파싱 헬퍼 ─────────────────────────────────────────────

def _cell_str(cell) -> str:
    return str(cell.value).strip() if cell.value is not None else ""

def _to_int(val: str) -> Optional[int]:
    if not val:
        return None
    try:
        return int(float(val.replace(",", "")))
    except (ValueError, AttributeError):
        return None

def _to_float(val: str) -> Optional[float]:
    if not val:
        return None
    try:
        return float(val.replace(",", ""))
    except (ValueError, AttributeError):
        return None

def _parse_date(val: str) -> str:
    """2026.02.26 → 2026-02-26, 이미 - 형식이면 그대로"""
    return val.replace(".", "-")


@app.post("/api/import/xlsx")
async def import_xlsx(file: UploadFile = File(...)):
    """
    마이클 xlsx 파일을 파싱하여 /api/import/preview 와 동일한 포맷으로 반환.
    DB에는 저장하지 않는다. confirm은 기존 /api/import/confirm 을 사용한다.
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"xlsx 파일을 열 수 없습니다: {e}")

    fuel_records: List[ImportRecordFuel] = []
    maint_records: List[ImportRecordMaintenance] = []
    other_records: List[ImportRecordOther] = []
    errors: List[str] = []

    # 주유(충전) 시트
    if "주유(충전)" in wb.sheetnames:
        ws = wb["주유(충전)"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            try:
                cols = [str(c).strip() if c is not None else "" for c in row]
                raw_date = cols[0]
                ftype    = cols[1] if cols[1] else "가득주유"
                amount   = _to_int(cols[2]) or 0
                unit_price = _to_int(cols[3])
                liters   = _to_float(cols[4])
                odometer = _to_int(cols[6]) or 0
                location = cols[7] or None
                memo     = cols[10] if len(cols) > 10 else ""
                parsed_date = _parse_date(raw_date)
                date.fromisoformat(parsed_date)  # 날짜 유효성 검사
                fuel_records.append(ImportRecordFuel(
                    date=parsed_date, type=ftype, amount=amount,
                    unit_price=unit_price, liters=liters, odometer=odometer,
                    location=location, memo=memo,
                ))
            except ValueError:
                errors.append(f"주유 {i}행: 날짜 형식 오류 ({row[0]!r})")
            except Exception as e:
                errors.append(f"주유 {i}행: {e}")

    # 정비 시트
    if "정비" in wb.sheetnames:
        ws = wb["정비"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            try:
                cols = [str(c).strip() if c is not None else "" for c in row]
                parsed_date = _parse_date(cols[0])
                date.fromisoformat(parsed_date)
                item     = cols[2] if len(cols) > 2 else ""
                amount   = _to_int(cols[3]) or 0
                odometer = _to_int(cols[4]) or 0
                location = cols[5] if len(cols) > 5 else None
                memo     = cols[6] if len(cols) > 6 else ""
                maint_records.append(ImportRecordMaintenance(
                    date=parsed_date, item=item, amount=amount,
                    odometer=odometer, location=location or None, memo=memo,
                ))
            except ValueError:
                errors.append(f"정비 {i}행: 날짜 형식 오류 ({row[0]!r})")
            except Exception as e:
                errors.append(f"정비 {i}행: {e}")

    # 기타 시트
    if "기타" in wb.sheetnames:
        ws = wb["기타"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            try:
                cols = [str(c).strip() if c is not None else "" for c in row]
                parsed_date = _parse_date(cols[0])
                date.fromisoformat(parsed_date)
                item     = cols[2] if len(cols) > 2 else ""
                amount   = _to_int(cols[3]) or 0
                odometer = _to_int(cols[4])
                location = cols[5] if len(cols) > 5 else None
                memo     = cols[6] if len(cols) > 6 else ""
                other_records.append(ImportRecordOther(
                    date=parsed_date, item=item, amount=amount,
                    odometer=odometer, location=location or None, memo=memo,
                ))
            except ValueError:
                errors.append(f"기타 {i}행: 날짜 형식 오류 ({row[0]!r})")
            except Exception as e:
                errors.append(f"기타 {i}행: {e}")

    wb.close()

    return {
        "counts": {
            "fuel":        len(fuel_records),
            "maintenance": len(maint_records),
            "other":       len(other_records),
        },
        "records": {
            "fuel":        [r.model_dump() for r in fuel_records],
            "maintenance": [r.model_dump() for r in maint_records],
            "other":       [r.model_dump() for r in other_records],
        },
        "vehicle": None,
        "errors": errors,
    }


@app.post("/api/import/preview")
def import_preview(body: ImportBody):
    """
    업로드된 JSON을 파싱·검증 후 미리보기 데이터 반환.
    DB에는 저장하지 않는다. 이상이 없으면 /api/import/confirm 을 호출할 것.
    """
    errors: List[str] = []

    # 날짜 형식 검증 헬퍼
    def check_date(val: str, label: str) -> bool:
        try:
            date.fromisoformat(val)
            return True
        except ValueError:
            errors.append(f"{label}: 날짜 형식 오류 ({val!r})")
            return False

    for i, r in enumerate(body.fuel):
        check_date(r.date, f"주유 {i+1}번")

    for i, r in enumerate(body.maintenance):
        check_date(r.date, f"정비 {i+1}번")

    for i, r in enumerate(body.other):
        check_date(r.date, f"기타 {i+1}번")

    return {
        "counts": {
            "fuel":        len(body.fuel),
            "maintenance": len(body.maintenance),
            "other":       len(body.other),
        },
        "records": {
            "fuel":        [r.model_dump() for r in body.fuel],
            "maintenance": [r.model_dump() for r in body.maintenance],
            "other":       [r.model_dump() for r in body.other],
        },
        "vehicle": body.vehicle.model_dump() if body.vehicle else None,
        "errors": errors,
    }


@app.post("/api/import/confirm", status_code=201)
def import_confirm(body: ImportBody):
    """
    미리보기에서 확인한 데이터를 실제 DB에 저장한다.
    fuel 레코드는 날짜 오름차순으로 삽입하여 연비를 올바르게 계산한다.
    """
    conn = get_db()

    # 차량 정보가 있으면 settings 에 반영 (빈 값은 덮어쓰지 않음)
    if body.vehicle:
        v = body.vehicle
        vehicle_fields = {
            "car_birth":  v.car_birth,
            "car_type":   v.car_type,
            "car_brand":  v.car_brand,
            "car_model":  v.car_model,
            "car_plate":  v.car_plate,
            "car_fuel":   v.car_fuel,
        }
        for key, value in vehicle_fields.items():
            if value:
                conn.execute(
                    "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
                    (key, value),
                )

    # 주유 — 날짜순 삽입으로 연비 자동 계산
    for record in sorted(body.fuel, key=lambda r: r.date):
        last = conn.execute(
            "SELECT odometer, liters FROM fuel ORDER BY date DESC, id DESC LIMIT 1"
        ).fetchone()
        interval_km = None
        fuel_economy = None
        if last and record.odometer > 0:
            diff = record.odometer - last["odometer"]
            if diff > 0:
                interval_km = diff
                if record.liters and record.liters > 0:
                    fuel_economy = round(interval_km / record.liters, 2)
        conn.execute(
            "INSERT INTO fuel (date, type, amount, unit_price, liters, odometer, interval_km, fuel_economy, location, memo)"
            " VALUES (?,?,?,?,?,?,?,?,?,?)",
            (record.date, record.type, record.amount, record.unit_price,
             record.liters, record.odometer, interval_km, fuel_economy, record.location, record.memo),
        )

    # 정비
    for record in body.maintenance:
        conn.execute(
            "INSERT INTO maintenance (date, category, item, amount, odometer, location, memo) VALUES (?,?,?,?,?,?,?)",
            (record.date, "정비", record.item, record.amount, record.odometer, record.location, record.memo),
        )

    # 기타
    for record in body.other:
        conn.execute(
            "INSERT INTO other (date, category, item, amount, odometer, location, memo) VALUES (?,?,?,?,?,?,?)",
            (record.date, "기타", record.item, record.amount, record.odometer, record.location, record.memo),
        )

    conn.commit()
    conn.close()

    return {
        "imported": {
            "fuel":        len(body.fuel),
            "maintenance": len(body.maintenance),
            "other":       len(body.other),
        }
    }


def seed_demo():
    if os.environ.get("DEMO_MODE", "").lower() != "true":
        return
    seed_path = BASE_DIR / "seed_demo.json"
    if not seed_path.exists():
        return
    conn = get_db()
    has_data = conn.execute("SELECT COUNT(*) FROM fuel").fetchone()[0] > 0
    conn.close()
    if has_data:
        return
    with open(seed_path, encoding="utf-8") as f:
        raw = json.load(f)
    import_confirm(ImportBody(**raw))


# 내보내기
@app.get("/api/export")
def export_data():
    """현재 DB 전체를 차필 가져오기 형식과 동일한 JSON으로 내보낸다."""
    content = json.dumps(build_export_data(), ensure_ascii=False, indent=2)
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=chapil_export.json"},
    )


# ── 주유 ──────────────────────────────────────────────────────────────
@app.get("/api/fuel")
def fuel_list():
    try:
        cached = redis_client.get("fuel_list")
        if cached:
            return json.loads(cached)
    except Exception:
        pass

    conn = get_db()
    rows = conn.execute("SELECT * FROM fuel ORDER BY date DESC, id DESC").fetchall()
    conn.close()
    result = [dict(row) for row in rows]
    try:
        redis_client.setex("fuel_list", 60, json.dumps(result))
    except Exception:
        pass
    return result


# status_code=201: REST 관례상 리소스 생성 성공 시 201 Created를 반환한다.
@app.post("/api/fuel", status_code=201)
def fuel_create(body: FuelBody):
    conn = get_db()
    try: redis_client.delete("fuel_list")
    except Exception: pass
    last = conn.execute(
        "SELECT odometer, liters FROM fuel ORDER BY date DESC, id DESC LIMIT 1"
    ).fetchone()

    interval_km = None
    fuel_economy = None
    if last and body.odometer > 0:
        diff = body.odometer - last["odometer"]
        if diff > 0:
            interval_km = diff
            if body.liters and body.liters > 0:
                fuel_economy = round(interval_km / body.liters, 2)

    cursor = conn.execute(
        "INSERT INTO fuel (date, type, amount, unit_price, liters, odometer, interval_km, fuel_economy, location, memo)"
        " VALUES (?,?,?,?,?,?,?,?,?,?)",
        (body.date, body.type, body.amount, body.unit_price, body.liters,
         body.odometer, interval_km, fuel_economy, body.location, body.memo),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM fuel WHERE id=?", (cursor.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.put("/api/fuel/{record_id}")
def fuel_update(record_id: int, body: FuelBody):
    conn = get_db()
    if not conn.execute("SELECT id FROM fuel WHERE id=?", (record_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다.")

    conn.execute(
        "UPDATE fuel SET date=?, type=?, amount=?, unit_price=?, liters=?, odometer=?, location=?, memo=? WHERE id=?",
        (body.date, body.type, body.amount, body.unit_price, body.liters,
         body.odometer, body.location, body.memo, record_id),
    )

    # 수정된 기록의 앞뒤 레코드를 찾아 연비를 재계산한다.
    prev = conn.execute(
        "SELECT odometer, liters FROM fuel WHERE date <= ? AND id != ? ORDER BY date DESC, id DESC LIMIT 1",
        (body.date, record_id),
    ).fetchone()
    next_ = conn.execute(
        "SELECT id, odometer, liters FROM fuel WHERE date >= ? AND id != ? ORDER BY date ASC, id ASC LIMIT 1",
        (body.date, record_id),
    ).fetchone()

    interval_km = None
    fuel_economy = None
    if prev and body.odometer > 0:
        diff = body.odometer - prev["odometer"]
        if diff > 0:
            interval_km = diff
            if body.liters and body.liters > 0:
                fuel_economy = round(interval_km / body.liters, 2)
    conn.execute(
        "UPDATE fuel SET interval_km=?, fuel_economy=? WHERE id=?",
        (interval_km, fuel_economy, record_id),
    )

    # 다음 기록의 연비도 현재 수정값 기준으로 다시 계산
    if next_:
        next_interval = next_["odometer"] - body.odometer
        next_economy = None
        if next_["liters"] and next_["liters"] > 0 and next_interval > 0:
            next_economy = round(next_interval / next_["liters"], 2)
        conn.execute(
            "UPDATE fuel SET interval_km=?, fuel_economy=? WHERE id=?",
            (next_interval, next_economy, next_["id"]),
        )

    conn.commit()
    try: redis_client.delete("fuel_list")
    except Exception: pass
    row = conn.execute("SELECT * FROM fuel WHERE id=?", (record_id,)).fetchone()
    conn.close()
    return dict(row)


# status_code=204: 삭제 성공 시 본문 없이 204 No Content를 반환한다.
@app.delete("/api/fuel/{record_id}", status_code=204)
def fuel_delete(record_id: int):
    conn = get_db()
    conn.execute("DELETE FROM fuel WHERE id=?", (record_id,))
    conn.commit()
    conn.close()
    try: redis_client.delete("fuel_list")
    except Exception: pass


# ── 정비 ──────────────────────────────────────────────────────────────
# 주의: /api/maintenance/items 는 반드시 /api/maintenance/{record_id} 보다
# 먼저 정의해야 한다. 그렇지 않으면 "items"가 record_id로 잘못 매칭된다.
@app.get("/api/maintenance/items")
def maintenance_items():
    return MAINTENANCE_ITEMS


@app.get("/api/maintenance")
def maintenance_list():
    conn = get_db()
    rows = conn.execute("SELECT * FROM maintenance ORDER BY date DESC, id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/maintenance", status_code=201)
def maintenance_create(body: MaintenanceBody):
    conn = get_db()
    cursor = conn.execute(
        "INSERT INTO maintenance (date, category, item, amount, odometer, location, memo) VALUES (?,?,?,?,?,?,?)",
        (body.date, "정비", body.item, body.amount, body.odometer, body.location, body.memo),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM maintenance WHERE id=?", (cursor.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.put("/api/maintenance/{record_id}")
def maintenance_update(record_id: int, body: MaintenanceBody):
    conn = get_db()
    if not conn.execute("SELECT id FROM maintenance WHERE id=?", (record_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다.")
    conn.execute(
        "UPDATE maintenance SET date=?, item=?, amount=?, odometer=?, location=?, memo=? WHERE id=?",
        (body.date, body.item, body.amount, body.odometer, body.location, body.memo, record_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM maintenance WHERE id=?", (record_id,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/maintenance/{record_id}", status_code=204)
def maintenance_delete(record_id: int):
    conn = get_db()
    conn.execute("DELETE FROM maintenance WHERE id=?", (record_id,))
    conn.commit()
    conn.close()


# ── 기타 ──────────────────────────────────────────────────────────────
@app.get("/api/other/items")
def other_items():
    return OTHER_ITEMS


@app.get("/api/other")
def other_list():
    conn = get_db()
    rows = conn.execute("SELECT * FROM other ORDER BY date DESC, id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/other", status_code=201)
def other_create(body: OtherBody):
    conn = get_db()
    cursor = conn.execute(
        "INSERT INTO other (date, category, item, amount, odometer, location, memo) VALUES (?,?,?,?,?,?,?)",
        (body.date, "기타", body.item, body.amount, body.odometer, body.location, body.memo),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM other WHERE id=?", (cursor.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.put("/api/other/{record_id}")
def other_update(record_id: int, body: OtherBody):
    conn = get_db()
    if not conn.execute("SELECT id FROM other WHERE id=?", (record_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다.")
    conn.execute(
        "UPDATE other SET date=?, item=?, amount=?, odometer=?, location=?, memo=? WHERE id=?",
        (body.date, body.item, body.amount, body.odometer, body.location, body.memo, record_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM other WHERE id=?", (record_id,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/other/{record_id}", status_code=204)
def other_delete(record_id: int):
    conn = get_db()
    conn.execute("DELETE FROM other WHERE id=?", (record_id,))
    conn.commit()
    conn.close()


# ── React SPA 서빙 ────────────────────────────────────────────────────
# Docker 빌드 시 dist/ 가 존재할 때만 활성화된다.
# 로컬 개발 환경에서는 Vite dev server(5173)가 프론트를 담당하므로 이 블록은 실행되지 않는다.
#
# /{full_path:path} : 위에서 정의한 /api/* 라우트에 매칭되지 않은
#                     모든 경로를 여기서 받아 처리한다.
# - dist/ 안에 해당 파일이 실제로 있으면 그 파일을 반환 (JS, CSS, 아이콘 등)
# - 없으면 index.html을 반환 → React Router가 클라이언트 사이드에서 라우팅 처리
if DIST_DIR.exists():
    @app.get("/{full_path:path}")
    def serve_spa(full_path: str):
        target = DIST_DIR / full_path
        if target.exists() and target.is_file():
            return FileResponse(target)
        return FileResponse(DIST_DIR / "index.html")
